import openpyxl
import os
from datetime import datetime,date
file_name = raw_input("File name?(file and test.py should be under same dir) ")
print file_name

wb = openpyxl.load_workbook(file_name)
sheet_names = wb.get_sheet_names()
Testing_res = wb.get_sheet_by_name(sheet_names[2])
Truth_res = wb.get_sheet_by_name(sheet_names[0])

print Testing_res.title
print Truth_res.title

TestingResEventCol = 0
TruthResEventCol = 0
TestingResTimeCol = 0
TruthResTimeCol = 0

for i in range(1,Testing_res.max_column + 1):
    if Testing_res.cell(row = 2,column = i).value == "Event Type":
        TestingResEventCol = i
    elif Testing_res.cell(row = 2,column = i).value == "Timestamp":
        TestingResTimeCol = i

for i in range(1,Truth_res.max_column + 1):
    if Truth_res.cell(row = 2,column = i).value == "Event Type":
        TruthResEventCol = i
    elif Truth_res.cell(row = 2,column = i).value == "Timestamp":
        TruthResTimeCol = i

#print TestingResEventCol,TestingResTimeCol,TruthResEventCol,TruthResTimeCol

"""Go Through the result workbook"""
"""Res_Dict[ID] = [film name,time subtraction,correct category,predicted category]"""
Res_Dict = {}
for i in range(3,Testing_res.max_row + 1):
    if Testing_res.cell(row = i,column = 1).value == None:
        print("oops")
        break
    else:
        Test_time = Testing_res.cell(row = i,column = TestingResTimeCol).value
        GT_time = Truth_res.cell(row = i,column = TruthResTimeCol).value
        if Test_time < GT_time:
            Test_time,GT_time = GT_time,Test_time
        sub_time = (datetime.combine(date.today(),Test_time) - datetime.combine(date.today(),GT_time)).seconds
        print Test_time,GT_time,sub_time,"secs"
        Name = Testing_res.cell(row = i,column = 2).value
        TruthEvent = Truth_res.cell(row = i,column = TruthResEventCol).value
        TestEvent = Testing_res.cell(row = i,column = TestingResEventCol).value
        Res_Dict[Testing_res.cell(row = i,column = 1).value] = [Name,sub_time,TruthEvent,TestEvent]
print Res_Dict
        
"""Overall_Dict[category] = [test items,correct,items,accuracy]"""

Overall_Dict = {"Blockage (1)" : [0,0,0.0],
                "Spray (2)" : [0,0,0.0],
                "Defocus (3)" : [0,0,0.0],
                "Redirect (4)" : [0,0,0.0]}

for keys in Res_Dict:
    category = Res_Dict[keys][2]
    predicted = Res_Dict[keys][3]
    Overall_Dict[category][0] += 1
    if category == predicted:
        Overall_Dict[category][1] += 1

for keys in Overall_Dict:
    if Overall_Dict[keys][0] != 0:
        Overall_Dict[keys][2] = float(Overall_Dict[keys][1]) / Overall_Dict[keys][0]

#print Overall_Dict

TimeError = int(raw_input("Legal Time Error =  secs? "))
for keys in Res_Dict:
    if Res_Dict[keys][1] <= TimeError:
        print Res_Dict[keys][0]
